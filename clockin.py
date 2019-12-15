from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from os import getenv, path


def start_work():
    # set file name from environment variable
    file_name = getenv('TIMESHEET')
    if file_name is not None:
        if file_name == "":
            file_name = "timesheet.xlsx"
        elif file_name.split('.')[1] != 'xlsx':
            file_name = file_name.split('.')[0] + '.xlsx'
    else:
        file_name = "timesheet.xlsx"

    # start work timing
    print('Starting work at ' + datetime.now().strftime("%#I:%M %p") + '...')
    start_time = datetime.now()
    input('Press enter to end work... ')
    print('Work done at ' + datetime.now().strftime("%#I:%M %p"))

    # end work timing and calculate hours
    end_time = datetime.now()
    work_duration = end_time - start_time
    hours = round(work_duration.seconds/3600, 2)
    print('Hours: ' + str(hours))
    # get task info
    task = input('Enter task info... ')
    print("Task: " + task)

    # check if inputted file exists
    # if yes, write to that file - assume correct format
    if path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
        row_number = sheet.max_row
        last_date = sheet['A' + str(row_number)]
        todays_date = datetime.now().strftime("%x")
        # if last entry from same day, append new info to that
        if last_date.value == todays_date:
            sheet['B' + str(row_number)] = str(float(sheet['B' + str(row_number)].value) + hours)
            sheet['C' + str(row_number)] = sheet['C' + str(row_number)].value + '; ' + task
        # else add new row with current date
        else:
            sheet["A" + str(row_number + 1)] = datetime.now().strftime("%x")
            sheet["B" + str(row_number + 1)] = str(hours)
            sheet["C" + str(row_number + 1)] = task
        workbook.save(filename=file_name)
        print("Saved work session to existing file: " + file_name)

    # if no, write to new file with default name
    else:
        workbook = Workbook()
        sheet = workbook.active
        bold_font = Font(bold=True)
        sheet["A1"] = "Date"
        sheet["B1"] = "Hours"
        sheet["C1"] = "Task"
        sheet["A1"].font = bold_font
        sheet["B1"].font = bold_font
        sheet["C1"].font = bold_font
        sheet["A2"] = datetime.now().strftime("%x")
        sheet["B2"] = str(hours)
        sheet["C2"] = task
        workbook.save(filename=file_name)
        print("Saved work session to new file: " + file_name)


if __name__ == "__main__":
    start_work()
